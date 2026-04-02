"""
경성뎐 : game_data -> xlsx 역변환 스크립트

사용법:
  python json_to_xlsx.py
  python json_to_xlsx.py ../game/data/game_data.js
  python json_to_xlsx.py ../game/data/game_data.js script.generated.xlsx

입력:
  - window.GAME_DATA = {...}; 형태의 js 파일
  - 또는 순수 json 파일

출력:
  - SceneTable
  - DialogTable
  - ChoiceTable
  - BranchTable
  - EvidenceTable
"""

import json
import os
import re
import sys

try:
    import openpyxl
    from openpyxl import Workbook
    from copy import copy
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)


DEFAULT_INPUT = os.path.join(os.path.dirname(__file__), "../game/data/game_data.js")
DEFAULT_OUTPUT = os.path.join(os.path.dirname(__file__), "script.generated.xlsx")


def read_text(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def parse_game_data(text):
    trimmed = text.strip()
    if trimmed.startswith("window.GAME_DATA"):
        trimmed = re.sub(r"^window\.GAME_DATA\s*=\s*", "", trimmed, count=1)
        trimmed = re.sub(r";\s*$", "", trimmed)
    return json.loads(trimmed)


def normalize_value(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def ensure_sheet(wb, title):
    if title in wb.sheetnames:
        return wb[title]
    return wb.create_sheet(title)


def normalize_header(value):
    return (value or "").strip()


def copy_cell_style(src, dst):
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)


def copy_column_style(ws, source_col, target_col):
    max_row = max(ws.max_row, 2)
    for row_idx in range(1, max_row + 1):
        copy_cell_style(ws.cell(row=row_idx, column=source_col), ws.cell(row=row_idx, column=target_col))

    source_letter = openpyxl.utils.get_column_letter(source_col)
    target_letter = openpyxl.utils.get_column_letter(target_col)
    if source_letter in ws.column_dimensions:
        ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width


def ensure_headers(ws, headers, aliases=None):
    aliases = aliases or {}

    actual_headers = [normalize_header(ws.cell(row=1, column=col_idx).value) for col_idx in range(1, ws.max_column + 1)]

    for header in headers:
        alias_list = [header] + aliases.get(header, [])
        found_idx = None

        for col_idx in range(1, ws.max_column + 1):
            current = normalize_header(ws.cell(row=1, column=col_idx).value)
            if current in alias_list:
                found_idx = col_idx
                break

        if found_idx is None:
            target_idx = ws.max_column + 1
            source_col = max(1, target_idx - 1)
            if source_col != target_idx:
                copy_column_style(ws, source_col, target_idx)
            ws.cell(row=1, column=target_idx, value=header)
            actual_headers.append(header)
        else:
            ws.cell(row=1, column=found_idx, value=header)

    return [normalize_header(ws.cell(row=1, column=col_idx).value) for col_idx in range(1, ws.max_column + 1)]


def clear_sheet_values(ws, actual_headers):
    max_row = ws.max_row or 1
    for row_idx in range(2, max_row + 1):
        for col_idx in range(1, len(actual_headers) + 1):
            ws.cell(row=row_idx, column=col_idx, value=None)


def copy_row_style(ws, source_row, target_row, width):
    for col_idx in range(1, width + 1):
        src = ws.cell(row=source_row, column=col_idx)
        dst = ws.cell(row=target_row, column=col_idx)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)


def write_sheet(ws, headers, rows, aliases=None):
    actual_headers = ensure_headers(ws, headers, aliases=aliases)
    clear_sheet_values(ws, actual_headers)
    template_row = 2 if ws.max_row >= 2 else None
    for row_idx, row in enumerate(rows, start=2):
        if template_row and row_idx > ws.max_row:
            copy_row_style(ws, template_row, row_idx, len(actual_headers))
        for col_idx, header in enumerate(actual_headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(actual_headers))}{max(len(rows) + 1, 1)}"


def build_scene_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        rows.append({
            "SceneID": scene.get("id", scene_id),
            "Chapter": scene.get("chapter"),
            "Title": scene.get("title"),
            "Background": scene.get("background"),
            "Music": scene.get("music"),
            "NextScene": scene.get("next_scene"),
            "Effect": scene.get("effect"),
        })
    return rows


def build_dialog_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        for line in sorted(scene.get("dialogues", []), key=lambda x: x.get("order", 0)):
            condition = line.get("condition") or {}
            rows.append({
                "SceneID": scene_id,
                "Order": line.get("order"),
                "Label": line.get("label"),
                "Speaker": line.get("speaker"),
                "Text": line.get("text"),
                "Style": line.get("style"),
                "Portrait": line.get("portrait"),
                "ConditionKey": condition.get("flag_key"),
                "ConditionValue": normalize_value(condition.get("flag_value")),
            })
    return rows


def build_choice_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        for choice in sorted(scene.get("choices", []), key=lambda x: x.get("order", 0)):
            rows.append({
                "SceneID": scene_id,
                "Order": choice.get("order"),
                "Text": choice.get("text"),
                "FlagKey": choice.get("flag_key"),
                "FlagValue": normalize_value(choice.get("flag_value")),
                "NextScene": choice.get("next_scene"),
                "NextDialogue": choice.get("next_dialogue"),
            })
    return rows


def build_branch_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        for branch in sorted(scene.get("branches", []), key=lambda x: x.get("order", 0)):
            rows.append({
                "SceneID": scene_id,
                "Order": branch.get("order"),
                "FlagKey": branch.get("flag_key"),
                "FlagValue": normalize_value(branch.get("flag_value")),
                "NextScene": branch.get("next_scene"),
            })
    return rows


def build_evidence_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        for ev in scene.get("evidence", []):
            rows.append({
                "EvidenceID": ev.get("evidence_id"),
                "SceneId": scene_id,
                "Trigger": normalize_value(ev.get("trigger")),
                "Name": ev.get("name"),
                "Description": ev.get("description"),
                "Image": ev.get("image"),
            })
    rows.sort(key=lambda x: ((x.get("SceneId") or ""), (x.get("EvidenceID") or "")))
    return rows


def build_workbook(data, output_path):
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = Workbook()
        wb.active.title = "SceneTable"

    write_sheet(
        ensure_sheet(wb, "SceneTable"),
        ["SceneID", "Chapter", "Title", "Background", "Music", "NextScene", "Effect"],
        build_scene_rows(data),
    )
    write_sheet(
        ensure_sheet(wb, "DialogTable"),
        ["SceneID", "Order", "Label", "Speaker", "Text", "Style", "Portrait", "ConditionKey", "ConditionValue"],
        build_dialog_rows(data),
        aliases={"ConditionKey": ["CoinditionKey"]},
    )
    write_sheet(
        ensure_sheet(wb, "ChoiceTable"),
        ["SceneID", "Order", "Text", "FlagKey", "FlagValue", "NextScene", "NextDialogue"],
        build_choice_rows(data),
    )
    write_sheet(
        ensure_sheet(wb, "BranchTable"),
        ["SceneID", "Order", "FlagKey", "FlagValue", "NextScene"],
        build_branch_rows(data),
    )
    write_sheet(
        ensure_sheet(wb, "EvidenceTable"),
        ["EvidenceID", "SceneId", "Trigger", "Name", "Description", "Image"],
        build_evidence_rows(data),
    )

    return wb


def main():
    input_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT
    output_path = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT

    if not os.path.exists(input_path):
        print(f"입력 파일 없음: {input_path}")
        sys.exit(1)

    text = read_text(input_path)
    data = parse_game_data(text)
    wb = build_workbook(data, output_path)
    temp_output_path = output_path + ".tmp"
    wb.save(temp_output_path)
    os.replace(temp_output_path, output_path)

    print(f"완료: {output_path}")
    print(f"  Scene 수: {len(data.get('scenes', {}))}")


if __name__ == "__main__":
    main()

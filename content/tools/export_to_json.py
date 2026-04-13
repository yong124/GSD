"""
경성뎐 : 여급 실종사건
Excel → game_data.js 변환 스크립트

사용법:
  python export_to_json.py

결과:
  ../../game/data/game_data.js  (window.GAME_DATA = {...} 형태)
  서버 없이 index.html을 파일로 열어도 동작함

Excel 컬럼 규칙 (PascalCase):
  SceneTable / Scenes        : SceneID, Chapter, Title, Background, Music, Effect, GoalKicker, GoalText, EvidencePromptTitle, EvidencePromptHint
  DialogTable / Dialogues    : 새 구조 기준
  ChoiceTable / Choices      : 새 구조 기준
  BranchTable / Branches     : 새 구조 기준
  ConditionTable / Conditions
  ChoiceGroupTable / ChoiceGroups
  EvidenceTable / Evidence
  EvidenceCategoryTable / EvidenceCategories
  CharacterTable / Characters
  CharacterEmotionTable / CharacterEmotions
  InvestigationTable / Investigations
  QuestionTable / Questions : QuestionID, Title, Detail, SortOrder, Category, ResolutionType, VisibleConditionGroupIDs, StateConditionsJSON, RelatedEvidenceIDs, SolutionEvidenceIDs, SolutionMode, ContradictionPrompt, ContradictionStatement, SolvedStateID, ResolvedDetail, SuccessToast, FailureToast, RewardStateID, RewardValue, RewardMode
  StateDescriptorTable / StateDescriptors : DescriptorID, TargetStateID, MinValue, MaxValue, Label, Detail

무시 규칙:
  시트명이 $로 시작하면 export 대상에서 제외
  컬럼명이 $로 시작하면 해당 컬럼은 JSON에 포함하지 않음
"""

import json
import sys
import os

try:
    import openpyxl
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)

EXCEL_PATH  = os.path.join(os.path.dirname(__file__), "../data/script.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "../../game/data/game_data.js")


def normalize_condition_row(row):
    condition_type = row.get("ConditionType")
    target_id = row.get("ConditionTargetID")
    normalized_type = condition_type
    normalized_target_id = target_id

    if condition_type == "SongsoonTrust":
        normalized_type = "Trust"
        normalized_target_id = "Songsoon"
    elif condition_type == "ResonanceLevel":
        normalized_type = "GaugeValue"
        normalized_target_id = "Erosion"
    elif condition_type == "InvestigationScore":
        normalized_type = "GaugeValue"
        normalized_target_id = "Credibility"
    elif condition_type == "ReadRitualScore":
        normalized_type = "GaugeValue"
        normalized_target_id = "ReadRitualScore"
    elif condition_type == "StateValue":
        if target_id in {"ReadRitualScore", "SolvedQuestionCount"}:
            normalized_type = "GaugeValue"
            normalized_target_id = target_id
        elif target_id == "CalledEditor":
            normalized_type = "ChoiceSelected"
            normalized_target_id = "Ch5PathContactEditor" if row.get("ConditionValue") is True else "Ch5PathNoContact"
        else:
            normalized_type = "ChoiceSelected"
            normalized_target_id = {
                "TouchedRoomWall": "Ch3Room4TouchWall",
                "QuestionSolved_QIpangyuCall": "QuestionSolved_QIpangyuCall",
                "QuestionSolved_QIpangyuMadness": "QuestionSolved_QIpangyuMadness",
                "QuestionSolved_QSonggeumMissing": "QuestionSolved_QSonggeumMissing",
                "QuestionSolved_QSonggeumRunaway": "QuestionSolved_QSonggeumRunaway",
                "QuestionSolved_QRitualLead": "QuestionSolved_QRitualLead",
                "QuestionSolved_QRitualAccident": "QuestionSolved_QRitualAccident",
            }.get(target_id, target_id)

    return normalized_type, normalized_target_id


def is_ignored_name(value):
    return isinstance(value, str) and value.strip().startswith("$")


def cell_val(cell):
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v != "" else None
    return v


def read_sheet(ws):
    rows = list(ws.iter_rows())
    if not rows:
        return []
    headers = [cell_val(c) for c in rows[0]]
    result = []
    for row in rows[1:]:
        record = {}
        for i, cell in enumerate(row):
            if (
                i < len(headers)
                and headers[i] is not None
                and not is_ignored_name(headers[i])
            ):
                record[headers[i]] = cell_val(cell)
        if any(v is not None for v in record.values()):
            result.append(record)
    return result


def resolve_sheet(wb, *candidates):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    available = ", ".join(wb.sheetnames)
    wanted = ", ".join(candidates)
    print(f"시트 없음: {wanted} — 현재 시트: {available}")
    sys.exit(1)


def resolve_optional_sheet(wb, *candidates):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    return None


def build_game_data(wb):
    scenes_raw = read_sheet(resolve_sheet(wb, "SceneTable", "Scenes"))
    dialogues_raw = read_sheet(resolve_sheet(wb, "DialogTable", "Dialogues"))
    choices_raw = read_sheet(resolve_sheet(wb, "ChoiceTable", "Choices"))
    branches_raw = read_sheet(resolve_sheet(wb, "BranchTable", "Branches")) if "BranchTable" in wb.sheetnames or "Branches" in wb.sheetnames else []
    conditions_raw = read_sheet(resolve_sheet(wb, "ConditionTable", "Conditions")) if "ConditionTable" in wb.sheetnames or "Conditions" in wb.sheetnames else []
    choice_groups_raw = read_sheet(resolve_sheet(wb, "ChoiceGroupTable", "ChoiceGroups")) if "ChoiceGroupTable" in wb.sheetnames or "ChoiceGroups" in wb.sheetnames else []
    evidence_raw = read_sheet(resolve_sheet(wb, "EvidenceTable", "Evidence"))
    evidence_categories_raw = read_sheet(resolve_sheet(wb, "EvidenceCategoryTable", "EvidenceCategories")) if "EvidenceCategoryTable" in wb.sheetnames or "EvidenceCategories" in wb.sheetnames else []
    characters_raw = read_sheet(resolve_sheet(wb, "CharacterTable", "Characters")) if "CharacterTable" in wb.sheetnames or "Characters" in wb.sheetnames else []
    character_emotions_raw = read_sheet(resolve_sheet(wb, "CharacterEmotionTable", "CharacterEmotions")) if "CharacterEmotionTable" in wb.sheetnames or "CharacterEmotions" in wb.sheetnames else []
    investigations_raw = read_sheet(resolve_sheet(wb, "InvestigationTable", "Investigations")) if "InvestigationTable" in wb.sheetnames or "Investigations" in wb.sheetnames else []
    questions_raw = read_sheet(resolve_sheet(wb, "QuestionTable", "Questions")) if "QuestionTable" in wb.sheetnames or "Questions" in wb.sheetnames else []
    state_descriptors_raw = read_sheet(resolve_sheet(wb, "StateDescriptorTable", "StateDescriptors")) if "StateDescriptorTable" in wb.sheetnames or "StateDescriptors" in wb.sheetnames else []
    gauges_ws = resolve_optional_sheet(wb, "GaugeTable", "Gauges")
    gauge_states_ws = resolve_optional_sheet(wb, "GaugeStateTable", "GaugeStates")
    effects_ws = resolve_optional_sheet(wb, "EffectTable", "Effects")
    gauges_raw = read_sheet(gauges_ws) if gauges_ws else []
    gauge_states_raw = read_sheet(gauge_states_ws) if gauge_states_ws else []
    effects_raw = read_sheet(effects_ws) if effects_ws else []

    # SceneID 기준으로 씬 딕셔너리 구성
    scenes = {}
    for s in scenes_raw:
        sid = s["SceneID"]
        scenes[sid] = {
            "id":         sid,
            "chapter":    s.get("Chapter"),
            "title":      s.get("Title"),
            "background": s.get("Background"),
            "music":      s.get("Music"),
            "effect":     s.get("Effect"),
            "goal_kicker": s.get("GoalKicker"),
            "goal_text":  s.get("GoalText"),
            "evidence_prompt_title": s.get("EvidencePromptTitle"),
            "evidence_prompt_hint": s.get("EvidencePromptHint"),
            "branches":   [],
            "dialogues":  [],
            "choices":    [],
            "evidence":   [],
        }

    # Dialogues 삽입 후 Order 정렬
    for d in dialogues_raw:
        sid = d.get("SceneID")
        if sid and sid in scenes:
            entry = {
                "order":    d.get("Order") or 0,
                "text":     d.get("Text") or "",
                "style":    d.get("Style") or "normal",
            }
            optional_fields = {
                "dialog_id": d.get("DialogID"),
                "speaker_id": d.get("SpeakerID"),
                "emotion_type": d.get("EmotionType"),
                "standing_slot": d.get("StandingSlot"),
                "focus_type": d.get("FocusType"),
                "enter_motion": d.get("EnterMotion"),
                "exit_motion": d.get("ExitMotion"),
                "idle_motion": d.get("IdleMotion"),
                "fx_type": d.get("FxType"),
                "choice_group_id": d.get("ChoiceGroupID"),
                "next_dialog_id": d.get("NextDialogID"),
                "effect_group_id": d.get("EffectGroupID"),
            }
            for key, value in optional_fields.items():
                if value is not None:
                    entry[key] = value
            if d.get("ConditionGroupID"):
                entry["condition_group_id"] = d.get("ConditionGroupID")
            scenes[sid]["dialogues"].append(entry)
    for sid in scenes:
        scenes[sid]["dialogues"].sort(key=lambda x: x["order"])

    # Choices 삽입 후 Order 정렬
    for c in choices_raw:
        sid = c.get("SceneID")
        if sid and sid in scenes:
            entry = {
                "order":      c.get("Order") or 0,
                "text":       c.get("Text") or "",
            }
            if c.get("ChoiceID"):
                entry["choice_id"] = c.get("ChoiceID")
            if c.get("ChoiceGroupID"):
                entry["choice_group_id"] = c.get("ChoiceGroupID")
            if c.get("ConditionGroupID"):
                entry["condition_group_id"] = c.get("ConditionGroupID")
            if c.get("NextType"):
                entry["next_type"] = c.get("NextType")
            if c.get("NextID") is not None:
                entry["next_id"] = c.get("NextID")
            if c.get("EvidenceID"):
                entry["evidence_id"] = c.get("EvidenceID")
            if c.get("EffectGroupID"):
                entry["effect_group_id"] = c.get("EffectGroupID")
            scenes[sid]["choices"].append(entry)
    for sid in scenes:
        scenes[sid]["choices"].sort(key=lambda x: x["order"])

    # Branches 삽입 후 Order 정렬
    for b in branches_raw:
        sid = b.get("SceneID")
        if sid and sid in scenes:
            scenes[sid]["branches"].append({
                "branch_id":   b.get("BranchID"),
                "order":      b.get("Order") or 0,
                "next_scene": b.get("NextSceneID") or "",
                "condition_group_id": b.get("ConditionGroupID"),
            })
    for sid in scenes:
        scenes[sid]["branches"].sort(key=lambda x: x["order"])

    # Evidence 삽입
    for e in evidence_raw:
        sid = e.get("SceneId")  # Evidence 시트는 SceneId (소문자 d)
        if sid and sid in scenes:
            scenes[sid]["evidence"].append({
                "evidence_id": e.get("EvidenceID"),
                "trigger":     e.get("Trigger") or "auto",
                "name":        e.get("Name") or "",
                "description": e.get("Description") or "",
                "image":       e.get("Image"),
                "category_id": e.get("CategoryID"),
            })

    characters = {}
    for c in characters_raw:
        character_id = c.get("CharacterID")
        if not character_id:
            continue
        characters[character_id] = {
            "id": character_id,
            "display_name": c.get("DisplayName") or character_id,
            "default_emotion_type": c.get("DefaultEmotionType") or "Neutral",
            "default_image_path": c.get("DefaultImagePath"),
            "role_text": c.get("RoleText"),
            "notebook_summary1": c.get("NotebookSummary1"),
            "notebook_summary2": c.get("NotebookSummary2"),
        }

    character_emotions = {}
    for row in character_emotions_raw:
        character_id = row.get("CharacterID")
        emotion_type = row.get("EmotionType")
        image_path = row.get("ImagePath")
        if not character_id or not emotion_type or not image_path:
            continue
        if character_id not in character_emotions:
            character_emotions[character_id] = {}
        character_emotions[character_id][emotion_type] = image_path

    first_scene = scenes_raw[0]["SceneID"] if scenes_raw else None

    conditions = []
    for row in conditions_raw:
        condition_id = row.get("ConditionID")
        if not condition_id:
            continue
        condition_type, condition_target_id = normalize_condition_row(row)
        conditions.append({
            "condition_id": condition_id,
            "condition_group_id": row.get("ConditionGroupID"),
            "condition_type": condition_type,
            "condition_target_id": condition_target_id,
            "compare_type": row.get("CompareType"),
            "condition_value": row.get("ConditionValue"),
        })

    choice_groups = []
    for row in choice_groups_raw:
        choice_group_id = row.get("ChoiceGroupID")
        if not choice_group_id:
            continue
        choice_groups.append({
            "choice_group_id": choice_group_id,
            "type": row.get("Type"),
            "answer_type": row.get("AnswerType"),
            "condition_group_id": row.get("ConditionGroupID"),
            "max_selectable": row.get("MaxSelectable"),
            "default_dialog_id": row.get("DefaultDialogID"),
        })

    evidence_categories = []
    for row in evidence_categories_raw:
        category_id = row.get("CategoryID")
        if not category_id:
            continue
        evidence_categories.append({
            "category_id": category_id,
            "category_title": row.get("CategoryTitle") or "",
            "category_hint": row.get("CategoryHint") or "",
        })

    investigations = []
    for row in investigations_raw:
        investigation_id = row.get("InvestigationID")
        if not investigation_id:
            continue
        investigations.append({
            "investigation_id": investigation_id,
            "title": row.get("Title") or "",
            "hint": row.get("Hint") or "",
            "budget": row.get("Budget"),
            "choice_group_id": row.get("ChoiceGroupID"),
        })

    questions = []
    for q in questions_raw:
        question_id = q.get("QuestionID")
        if not question_id:
            continue
        state_conditions_raw = str(q.get("StateConditionsJSON") or "").strip()
        state_conditions = []
        if state_conditions_raw:
            try:
                parsed = json.loads(state_conditions_raw)
            except json.JSONDecodeError as exc:
                raise ValueError(f"QuestionID={question_id} StateConditionsJSON 파싱 실패: {exc}") from exc
            state_conditions = parsed if isinstance(parsed, list) else []
        questions.append({
            "question_id": question_id,
            "title": q.get("Title") or "",
            "detail": q.get("Detail") or "",
            "sort_order": q.get("SortOrder"),
            "category": q.get("Category"),
            "resolution_type": q.get("ResolutionType") or "Evidence",
            "visible_condition_group_ids": [part.strip() for part in str(q.get("VisibleConditionGroupIDs") or "").split(",") if part and part.strip()],
            "state_conditions": state_conditions,
            "related_evidence_ids": [part.strip() for part in str(q.get("RelatedEvidenceIDs") or "").split(",") if part and part.strip()],
            "solution_evidence_ids": [part.strip() for part in str(q.get("SolutionEvidenceIDs") or "").split(",") if part and part.strip()],
            "solution_mode": q.get("SolutionMode") or "",
            "contradiction_prompt": q.get("ContradictionPrompt") or "",
            "contradiction_statement": q.get("ContradictionStatement") or "",
            "solved_state_id": q.get("SolvedStateID") or "",
            "resolved_detail": q.get("ResolvedDetail") or "",
            "success_toast": q.get("SuccessToast") or "",
            "failure_toast": q.get("FailureToast") or "",
            "reward_state_id": q.get("RewardStateID") or "",
            "reward_value": q.get("RewardValue"),
            "reward_mode": q.get("RewardMode") or "",
        })

    state_descriptors = []
    for row in state_descriptors_raw:
        descriptor_id = row.get("DescriptorID")
        if not descriptor_id:
            continue
        state_descriptors.append({
            "descriptor_id": descriptor_id,
            "target_state_type": row.get("TargetStateType") or ("Derived" if row.get("TargetStateID") == "InvestigationProgress" else "Numeric"),
            "target_state_id": row.get("TargetStateID"),
            "min_value": row.get("MinValue"),
            "max_value": row.get("MaxValue"),
            "label": row.get("Label") or "",
            "detail": row.get("Detail") or "",
        })

    gauges = []
    for row in gauges_raw:
        gauge_id = row.get("GaugeID")
        if not gauge_id:
            continue
        gauges.append({
            "gauge_id": gauge_id,
            "label": row.get("Label") or "",
            "min_value": row.get("MinValue"),
            "max_value": row.get("MaxValue"),
            "default_value": row.get("DefaultValue"),
            "hud_visible": row.get("HudVisible"),
            "hud_order": row.get("HudOrder"),
        })

    gauge_states = []
    for row in gauge_states_raw:
        gauge_id = row.get("GaugeID")
        if not gauge_id:
            continue
        gauge_states.append({
            "gauge_id": gauge_id,
            "min_value": row.get("MinValue"),
            "max_value": row.get("MaxValue"),
            "label": row.get("Label") or "",
            "hud_color": row.get("HudColor") or "",
            "detail": row.get("Detail") or "",
            "trigger_scene_id": row.get("TriggerSceneID"),
        })

    effects = []
    for row in effects_raw:
        effect_group_id = row.get("EffectGroupID")
        if not effect_group_id:
            continue
        effects.append({
            "effect_group_id": effect_group_id,
            "effect_type": row.get("EffectType"),
            "gauge_id": row.get("GaugeID"),
            "gauge_delta": row.get("GaugeDelta"),
            "evidence_id": row.get("EvidenceID"),
            "trust_character_id": row.get("TrustCharacterID"),
            "trust_delta": row.get("TrustDelta"),
        })

    return {
        "first_scene": first_scene,
        "conditions": conditions,
        "choice_groups": choice_groups,
        "evidence_categories": evidence_categories,
        "characters": characters,
        "character_emotions": character_emotions,
        "investigations": investigations,
        "questions": questions,
        "state_descriptors": state_descriptors,
        "gauges": gauges,
        "gauge_states": gauge_states,
        "effects": effects,
        "scenes": scenes,
    }


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"파일 없음: {EXCEL_PATH}")
        sys.exit(1)

    print(f"읽는 중: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    ignored_sheets = [name for name in wb.sheetnames if is_ignored_name(name)]
    if ignored_sheets:
        print(f"무시된 시트: {', '.join(ignored_sheets)}")

    data = build_game_data(wb)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json_str = json.dumps(data, ensure_ascii=False, indent=2)
        f.write(f"window.GAME_DATA = {json_str};\n")

    scene_count = len(data["scenes"])
    print(f"완료: {OUTPUT_PATH}")
    print(f"  씬 {scene_count}개 변환됨 / 첫 씬: {data['first_scene']}")


if __name__ == "__main__":
    main()

"""
경성뎐 : game_data -> generated xlsx

원본 script.xlsx 양식/표는 건드리지 않고,
복붙용 generated xlsx와 보조 산출물을 생성한다.
"""

import argparse
import csv
import json
import os
import re
import shutil
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)


DEFAULT_INPUT = os.path.join(os.path.dirname(__file__), "../../game/data/game_data.js")
DEFAULT_OUTPUT = os.path.join(os.path.dirname(__file__), "../generated/script.generated.xlsx")

SHEET_DEFS = {
    "SceneTable": {
        "headers": ["SceneID", "Chapter", "Title", "Background", "Music", "NextScene", "Effect", "GoalKicker", "GoalText"],
    },
    "DialogTable": {
        "headers": ["SceneID", "Order", "Speaker", "SpeakerID", "EmotionType", "StandingSlot", "FocusType", "EnterMotion", "ExitMotion", "IdleMotion", "FxType", "Text", "Style", "Portrait", "ConditionKey", "ConditionValue", "Label"],
    },
    "ChoiceTable": {
        "headers": ["SceneID", "Order", "Text", "FlagKey", "FlagValue", "NextScene", "NextDialogue"],
    },
    "BranchTable": {
        "headers": ["SceneID", "FlagKey", "FlagValue", "NextScene", "Order"],
    },
    "EvidenceTable": {
        "headers": ["EvidenceID", "SceneId", "Trigger", "Name", "Description", "Image", "CategoryID", "CategoryTitle", "CategoryHint"],
    },
    "CharacterTable": {
        "headers": ["CharacterID", "DisplayName", "DefaultEmotionType", "DefaultImagePath", "RoleText", "NotebookSummary1", "NotebookSummary2"],
    },
    "CharacterEmotionTable": {
        "headers": ["CharacterID", "EmotionType", "ImagePath"],
    },
    "QuestionTable": {
        "headers": ["QuestionID", "Title", "Detail", "SortOrder", "Category", "VisibleRuleID", "StateRuleID", "RelatedEvidenceIDs", "SolutionEvidenceID", "SolutionEvidenceIDs", "SolutionMode", "SolvedFlagID", "ResolvedDetail", "SuccessToast", "FailureToast", "RewardFlagID", "RewardValue", "RewardMode"],
    },
    "StateDescriptorTable": {
        "headers": ["DescriptorID", "TargetFlagID", "MinValue", "MaxValue", "Label", "Detail"],
    },
    "RuleTable": {
        "headers": ["RuleRowID", "RuleID", "RuleKind", "FactType", "FactKey", "Operator", "Value", "ResultValue", "Priority"],
    },
}


def parse_args():
    parser = argparse.ArgumentParser(description="game_data.js를 generated xlsx로 변환")
    parser.add_argument("input", nargs="?", default=DEFAULT_INPUT, help="입력 game_data.js 경로")
    parser.add_argument("output", nargs="?", default=DEFAULT_OUTPUT, help="출력 xlsx 경로")
    parser.add_argument(
        "--sheets",
        default="all",
        help="생성할 시트 목록. 예: SceneTable,DialogTable 또는 all",
    )
    parser.add_argument(
        "--with-delimited",
        action="store_true",
        help="시트별 csv/tsv 파일도 함께 생성",
    )
    return parser.parse_args()


def read_text(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def parse_game_data(text):
    trimmed = text.strip()
    if trimmed.startswith("window.GAME_DATA"):
        trimmed = re.sub(r"^window\.GAME_DATA\s*=\s*", "", trimmed, count=1)
        trimmed = re.sub(r";\s*$", "", trimmed)
    return json.loads(trimmed)


def normalize_value(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def append_sheet(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    ws.freeze_panes = "A2"


def build_scene_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        rows.append({
            "SceneID": scene.get("id", scene_id),
            "Chapter": scene.get("chapter"),
            "Title": scene.get("title"),
            "Background": scene.get("background"),
            "Music": scene.get("music"),
            "NextScene": scene.get("next_scene"),
            "Effect": scene.get("effect"),
            "GoalKicker": scene.get("goal_kicker"),
            "GoalText": scene.get("goal_text"),
        })
    return rows


def build_dialog_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        for line in sorted(scene.get("dialogues", []), key=lambda x: x.get("order", 0)):
            condition = line.get("condition") or {}
            rows.append({
                "SceneID": scene_id,
                "Order": line.get("order"),
                "Speaker": line.get("speaker"),
                "SpeakerID": line.get("speaker_id"),
                "EmotionType": line.get("emotion_type"),
                "StandingSlot": line.get("standing_slot"),
                "FocusType": line.get("focus_type"),
                "EnterMotion": line.get("enter_motion"),
                "ExitMotion": line.get("exit_motion"),
                "IdleMotion": line.get("idle_motion"),
                "FxType": line.get("fx_type"),
                "Text": line.get("text"),
                "Style": line.get("style"),
                "Portrait": line.get("portrait"),
                "ConditionKey": condition.get("flag_key"),
                "ConditionValue": normalize_value(condition.get("flag_value")),
                "Label": line.get("label"),
            })
    return rows


def build_choice_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        for choice in sorted(scene.get("choices", []), key=lambda x: x.get("order", 0)):
            rows.append({
                "SceneID": scene_id,
                "Order": choice.get("order"),
                "Text": choice.get("text"),
                "FlagKey": choice.get("flag_key"),
                "FlagValue": normalize_value(choice.get("flag_value")),
                "NextScene": choice.get("next_scene"),
                "NextDialogue": choice.get("next_dialogue"),
            })
    return rows


def build_branch_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        for branch in sorted(scene.get("branches", []), key=lambda x: x.get("order", 0)):
            rows.append({
                "SceneID": scene_id,
                "FlagKey": branch.get("flag_key"),
                "FlagValue": normalize_value(branch.get("flag_value")),
                "NextScene": branch.get("next_scene"),
                "Order": branch.get("order"),
            })
    return rows


def build_evidence_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        for ev in scene.get("evidence", []):
            rows.append({
                "EvidenceID": ev.get("evidence_id"),
                "SceneId": scene_id,
                "Trigger": normalize_value(ev.get("trigger")),
                "Name": ev.get("name"),
                "Description": ev.get("description"),
                "Image": ev.get("image"),
                "CategoryID": ev.get("category_id"),
                "CategoryTitle": ev.get("category_title"),
                "CategoryHint": ev.get("category_hint"),
            })
    rows.sort(key=lambda x: ((x.get("SceneId") or ""), (x.get("EvidenceID") or "")))
    return rows


def build_character_rows(data):
    rows = []
    characters = data.get("characters", {})
    for character_id in sorted(characters.keys()):
        character = characters[character_id]
        rows.append({
            "CharacterID": character.get("id", character_id),
            "DisplayName": character.get("display_name"),
            "DefaultEmotionType": character.get("default_emotion_type"),
            "DefaultImagePath": character.get("default_image_path"),
            "RoleText": character.get("role_text"),
            "NotebookSummary1": character.get("notebook_summary1"),
            "NotebookSummary2": character.get("notebook_summary2"),
        })
    return rows


def build_character_emotion_rows(data):
    rows = []
    character_emotions = data.get("character_emotions", {})
    for character_id in sorted(character_emotions.keys()):
        emotions = character_emotions[character_id] or {}
        for emotion_type in sorted(emotions.keys()):
            rows.append({
                "CharacterID": character_id,
                "EmotionType": emotion_type,
                "ImagePath": emotions.get(emotion_type),
            })
    return rows


def build_question_rows(data):
    rows = []
    for question in data.get("questions", []) or []:
        rows.append({
            "QuestionID": question.get("question_id"),
            "Title": question.get("title"),
            "Detail": question.get("detail"),
            "SortOrder": question.get("sort_order"),
            "Category": question.get("category"),
            "VisibleRuleID": question.get("visible_rule_id"),
            "StateRuleID": question.get("state_rule_id"),
            "RelatedEvidenceIDs": ", ".join(question.get("related_evidence_ids", []) or []),
            "SolutionEvidenceID": question.get("solution_evidence_id"),
            "SolutionEvidenceIDs": ", ".join(question.get("solution_evidence_ids", []) or []),
            "SolutionMode": question.get("solution_mode"),
            "SolvedFlagID": question.get("solved_flag_id"),
            "ResolvedDetail": question.get("resolved_detail"),
            "SuccessToast": question.get("success_toast"),
            "FailureToast": question.get("failure_toast"),
            "RewardFlagID": question.get("reward_flag_id"),
            "RewardValue": question.get("reward_value"),
            "RewardMode": question.get("reward_mode"),
        })
    rows.sort(key=lambda x: ((x.get("SortOrder") or 0), (x.get("QuestionID") or "")))
    return rows


def build_state_descriptor_rows(data):
    rows = []
    for descriptor in data.get("state_descriptors", []) or []:
        rows.append({
            "DescriptorID": descriptor.get("descriptor_id"),
            "TargetFlagID": descriptor.get("target_flag_id"),
            "MinValue": descriptor.get("min_value"),
            "MaxValue": descriptor.get("max_value"),
            "Label": descriptor.get("label"),
            "Detail": descriptor.get("detail"),
        })
    rows.sort(key=lambda x: (x.get("TargetFlagID") or "", x.get("MinValue") if x.get("MinValue") is not None else 0, x.get("DescriptorID") or ""))
    return rows


def build_rule_rows(data):
    rows = []
    for rule in data.get("rules", []) or []:
        rows.append({
            "RuleRowID": rule.get("rule_row_id"),
            "RuleID": rule.get("rule_id"),
            "RuleKind": rule.get("rule_kind"),
            "FactType": rule.get("fact_type"),
            "FactKey": rule.get("fact_key"),
            "Operator": rule.get("operator"),
            "Value": rule.get("value"),
            "ResultValue": rule.get("result_value"),
            "Priority": rule.get("priority"),
        })
    rows.sort(key=lambda x: ((x.get("RuleID") or ""), (x.get("Priority") if x.get("Priority") is not None else 0), (x.get("RuleRowID") or "")))
    return rows


def build_sheet_rows(data):
    return {
        "SceneTable": build_scene_rows(data),
        "DialogTable": build_dialog_rows(data),
        "ChoiceTable": build_choice_rows(data),
        "BranchTable": build_branch_rows(data),
        "EvidenceTable": build_evidence_rows(data),
        "CharacterTable": build_character_rows(data),
        "CharacterEmotionTable": build_character_emotion_rows(data),
        "QuestionTable": build_question_rows(data),
        "StateDescriptorTable": build_state_descriptor_rows(data),
        "RuleTable": build_rule_rows(data),
    }


def resolve_selected_sheets(sheet_arg):
    if sheet_arg == "all":
        return list(SHEET_DEFS.keys())

    requested = [name.strip() for name in sheet_arg.split(",") if name.strip()]
    invalid = [name for name in requested if name not in SHEET_DEFS]
    if invalid:
        print(f"알 수 없는 시트명: {', '.join(invalid)}")
        sys.exit(1)
    return requested


def build_workbook(data, selected_sheets, input_path):
    wb = Workbook()
    wb.remove(wb.active)

    sheet_rows = build_sheet_rows(data)
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    meta_ws = wb.create_sheet("Meta")
    meta_ws.append(["Key", "Value"])
    meta_ws.append(["GeneratedAt", created_at])
    meta_ws.append(["Source", os.path.abspath(input_path)])
    meta_ws.append(["FirstScene", data.get("first_scene")])
    meta_ws.append(["SelectedSheets", ", ".join(selected_sheets)])
    meta_ws.append(["SceneCount", len(data.get("scenes", {}))])
    meta_ws.freeze_panes = "A2"

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Sheet", "RowCount"])
    for sheet_name in selected_sheets:
        summary_ws.append([sheet_name, len(sheet_rows[sheet_name])])
    summary_ws.freeze_panes = "A2"

    for sheet_name in selected_sheets:
        ws = wb.create_sheet(sheet_name)
        append_sheet(ws, SHEET_DEFS[sheet_name]["headers"], sheet_rows[sheet_name])

    return wb, sheet_rows, created_at


def write_delimited_files(output_path, selected_sheets, sheet_rows):
    base_dir = os.path.splitext(output_path)[0] + "_delimited"
    shutil.rmtree(base_dir, ignore_errors=True)
    os.makedirs(base_dir, exist_ok=True)

    for sheet_name in selected_sheets:
        headers = SHEET_DEFS[sheet_name]["headers"]
        rows = sheet_rows[sheet_name]

        csv_path = os.path.join(base_dir, f"{sheet_name}.csv")
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow({header: row.get(header) for header in headers})

        tsv_path = os.path.join(base_dir, f"{sheet_name}.tsv")
        with open(tsv_path, "w", encoding="utf-8-sig", newline="") as tsv_file:
            writer = csv.DictWriter(tsv_file, fieldnames=headers, delimiter="\t")
            writer.writeheader()
            for row in rows:
                writer.writerow({header: row.get(header) for header in headers})

    return base_dir


def main():
    args = parse_args()
    input_path = args.input
    output_path = args.output

    if not os.path.exists(input_path):
        print(f"입력 파일 없음: {input_path}")
        sys.exit(1)

    selected_sheets = resolve_selected_sheets(args.sheets)
    data = parse_game_data(read_text(input_path))
    wb, sheet_rows, created_at = build_workbook(data, selected_sheets, input_path)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    print(f"완료: {output_path}")
    print(f"  생성 시각: {created_at}")
    print(f"  첫 씬: {data.get('first_scene')}")
    print(f"  포함 시트: {', '.join(selected_sheets)}")
    for sheet_name in selected_sheets:
        print(f"  - {sheet_name}: {len(sheet_rows[sheet_name])}행")

    if args.with_delimited:
        delimited_dir = write_delimited_files(output_path, selected_sheets, sheet_rows)
        print(f"  보조 파일: {delimited_dir}")


if __name__ == "__main__":
    main()

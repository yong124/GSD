"""
경성뎐 : game_data -> generated xlsx

원본 script.xlsx 양식/표는 건드리지 않고,
복붙용 generated xlsx를 새로 생성한다.
"""

import json
import os
import re
import sys

try:
    from openpyxl import Workbook
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)


DEFAULT_INPUT = os.path.join(os.path.dirname(__file__), "../game/data/game_data.js")
DEFAULT_OUTPUT = os.path.join(os.path.dirname(__file__), "script.generated.xlsx")


def read_text(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def parse_game_data(text):
    trimmed = text.strip()
    if trimmed.startswith("window.GAME_DATA"):
        trimmed = re.sub(r"^window\.GAME_DATA\s*=\s*", "", trimmed, count=1)
        trimmed = re.sub(r";\s*$", "", trimmed)
    return json.loads(trimmed)


def normalize_value(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def append_sheet(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    ws.freeze_panes = "A2"


def build_scene_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        rows.append({
            "SceneID": scene.get("id", scene_id),
            "Chapter": scene.get("chapter"),
            "Title": scene.get("title"),
            "Background": scene.get("background"),
            "Music": scene.get("music"),
            "NextScene": scene.get("next_scene"),
            "Effect": scene.get("effect"),
        })
    return rows


def build_dialog_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        for line in sorted(scene.get("dialogues", []), key=lambda x: x.get("order", 0)):
            condition = line.get("condition") or {}
            rows.append({
                "SceneID": scene_id,
                "Order": line.get("order"),
                "Speaker": line.get("speaker"),
                "Text": line.get("text"),
                "Style": line.get("style"),
                "Portrait": line.get("portrait"),
                "ConditionKey": condition.get("flag_key"),
                "ConditionValue": normalize_value(condition.get("flag_value")),
                "Label": line.get("label"),
            })
    return rows


def build_choice_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        for choice in sorted(scene.get("choices", []), key=lambda x: x.get("order", 0)):
            rows.append({
                "SceneID": scene_id,
                "Order": choice.get("order"),
                "Text": choice.get("text"),
                "FlagKey": choice.get("flag_key"),
                "FlagValue": normalize_value(choice.get("flag_value")),
                "NextScene": choice.get("next_scene"),
                "NextDialogue": choice.get("next_dialogue"),
            })
    return rows


def build_branch_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        for branch in sorted(scene.get("branches", []), key=lambda x: x.get("order", 0)):
            rows.append({
                "SceneID": scene_id,
                "FlagKey": branch.get("flag_key"),
                "FlagValue": normalize_value(branch.get("flag_value")),
                "NextScene": branch.get("next_scene"),
                "Order": branch.get("order"),
            })
    return rows


def build_evidence_rows(data):
    rows = []
    scenes = data.get("scenes", {})
    for scene_id in sorted(scenes.keys()):
        scene = scenes[scene_id]
        for ev in scene.get("evidence", []):
            rows.append({
                "EvidenceID": ev.get("evidence_id"),
                "SceneId": scene_id,
                "Trigger": normalize_value(ev.get("trigger")),
                "Name": ev.get("name"),
                "Description": ev.get("description"),
                "Image": ev.get("image"),
            })
    rows.sort(key=lambda x: ((x.get("SceneId") or ""), (x.get("EvidenceID") or "")))
    return rows


def build_workbook(data):
    wb = Workbook()

    ws = wb.active
    ws.title = "SceneTable"
    append_sheet(
        ws,
        ["SceneID", "Chapter", "Title", "Background", "Music", "NextScene", "Effect"],
        build_scene_rows(data),
    )

    ws = wb.create_sheet("DialogTable")
    append_sheet(
        ws,
        ["SceneID", "Order", "Speaker", "Text", "Style", "Portrait", "ConditionKey", "ConditionValue", "Label"],
        build_dialog_rows(data),
    )

    ws = wb.create_sheet("ChoiceTable")
    append_sheet(
        ws,
        ["SceneID", "Order", "Text", "FlagKey", "FlagValue", "NextScene", "NextDialogue"],
        build_choice_rows(data),
    )

    ws = wb.create_sheet("BranchTable")
    append_sheet(
        ws,
        ["SceneID", "FlagKey", "FlagValue", "NextScene", "Order"],
        build_branch_rows(data),
    )

    ws = wb.create_sheet("EvidenceTable")
    append_sheet(
        ws,
        ["EvidenceID", "SceneId", "Trigger", "Name", "Description", "Image"],
        build_evidence_rows(data),
    )

    return wb


def main():
    input_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT
    output_path = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT

    if not os.path.exists(input_path):
        print(f"입력 파일 없음: {input_path}")
        sys.exit(1)

    data = parse_game_data(read_text(input_path))
    wb = build_workbook(data)
    wb.save(output_path)

    print(f"완료: {output_path}")
    print(f"  Scene 수: {len(data.get('scenes', {}))}")


if __name__ == "__main__":
    main()

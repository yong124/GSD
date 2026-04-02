"""
경성뎐 : 여급 실종사건
Excel → game_data.js 변환 스크립트

사용법:
  python export_to_json.py

결과:
  ../game/data/game_data.js  (window.GAME_DATA = {...} 형태)
  서버 없이 index.html을 파일로 열어도 동작함

Excel 컬럼 규칙 (PascalCase):
  SceneTable / Scenes        : SceneID, Chapter, Title, Background, Music, NextScene, Effect
  DialogTable / Dialogues    : SceneID, Order, Label, Speaker, Text, Style, Portrait, ConditionKey, ConditionValue
  ChoiceTable / Choices      : SceneID, Order, Text, FlagKey, FlagValue, NextScene, NextDialogue
  BranchTable / Branches     : SceneID, Order, FlagKey, FlagValue, NextScene
  EvidenceTable / Evidence   : EvidenceID, SceneId, Trigger, Name, Description, Image

무시 규칙:
  시트명이 $로 시작하면 export 대상에서 제외
  컬럼명이 $로 시작하면 해당 컬럼은 JSON에 포함하지 않음
"""

import json
import sys
import os

try:
    import openpyxl
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)

EXCEL_PATH  = os.path.join(os.path.dirname(__file__), "script.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "../game/data/game_data.js")


def is_ignored_name(value):
    return isinstance(value, str) and value.strip().startswith("$")


def cell_val(cell):
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v != "" else None
    return v


def read_sheet(ws):
    rows = list(ws.iter_rows())
    if not rows:
        return []
    headers = [cell_val(c) for c in rows[0]]
    result = []
    for row in rows[1:]:
        record = {}
        for i, cell in enumerate(row):
            if (
                i < len(headers)
                and headers[i] is not None
                and not is_ignored_name(headers[i])
            ):
                record[headers[i]] = cell_val(cell)
        if any(v is not None for v in record.values()):
            result.append(record)
    return result


def resolve_sheet(wb, *candidates):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    available = ", ".join(wb.sheetnames)
    wanted = ", ".join(candidates)
    print(f"시트 없음: {wanted} — 현재 시트: {available}")
    sys.exit(1)


def build_game_data(wb):
    scenes_raw = read_sheet(resolve_sheet(wb, "SceneTable", "Scenes"))
    dialogues_raw = read_sheet(resolve_sheet(wb, "DialogTable", "Dialogues"))
    choices_raw = read_sheet(resolve_sheet(wb, "ChoiceTable", "Choices"))
    branches_raw = read_sheet(resolve_sheet(wb, "BranchTable", "Branches")) if "BranchTable" in wb.sheetnames or "Branches" in wb.sheetnames else []
    evidence_raw = read_sheet(resolve_sheet(wb, "EvidenceTable", "Evidence"))

    # SceneID 기준으로 씬 딕셔너리 구성
    scenes = {}
    for s in scenes_raw:
        sid = s["SceneID"]
        scenes[sid] = {
            "id":         sid,
            "chapter":    s.get("Chapter"),
            "title":      s.get("Title"),
            "background": s.get("Background"),
            "music":      s.get("Music"),
            "next_scene": s.get("NextScene"),
            "effect":     s.get("Effect"),
            "branches":   [],
            "dialogues":  [],
            "choices":    [],
            "evidence":   [],
        }

    # Dialogues 삽입 후 Order 정렬
    for d in dialogues_raw:
        sid = d.get("SceneID")
        if sid and sid in scenes:
            entry = {
                "order":    d.get("Order") or 0,
                "speaker":  d.get("Speaker") or "",
                "text":     d.get("Text") or "",
                "style":    d.get("Style") or "normal",
                "portrait": d.get("Portrait"),
                "condition": None,
            }
            if d.get("Label"):
                entry["label"] = d["Label"]
            cond_key = d.get("ConditionKey")
            cond_val = d.get("ConditionValue")
            if cond_key:
                entry["condition"] = {"flag_key": cond_key, "flag_value": cond_val}
            scenes[sid]["dialogues"].append(entry)
    for sid in scenes:
        scenes[sid]["dialogues"].sort(key=lambda x: x["order"])

    # Choices 삽입 후 Order 정렬
    for c in choices_raw:
        sid = c.get("SceneID")
        if sid and sid in scenes:
            entry = {
                "order":      c.get("Order") or 0,
                "text":       c.get("Text") or "",
                "flag_key":   c.get("FlagKey"),
                "flag_value": c.get("FlagValue"),
                "next_scene": c.get("NextScene"),
            }
            if c.get("NextDialogue"):
                entry["next_dialogue"] = c["NextDialogue"]
            scenes[sid]["choices"].append(entry)
    for sid in scenes:
        scenes[sid]["choices"].sort(key=lambda x: x["order"])

    # Branches 삽입 후 Order 정렬
    for b in branches_raw:
        sid = b.get("SceneID")
        if sid and sid in scenes:
            scenes[sid]["branches"].append({
                "order":      b.get("Order") or 0,
                "flag_key":   b.get("FlagKey") or "",
                "flag_value": b.get("FlagValue"),
                "next_scene": b.get("NextScene") or "",
            })
    for sid in scenes:
        scenes[sid]["branches"].sort(key=lambda x: x["order"])

    # Evidence 삽입
    for e in evidence_raw:
        sid = e.get("SceneId")  # Evidence 시트는 SceneId (소문자 d)
        if sid and sid in scenes:
            scenes[sid]["evidence"].append({
                "evidence_id": e.get("EvidenceID"),
                "trigger":     e.get("Trigger") or "auto",
                "name":        e.get("Name") or "",
                "description": e.get("Description") or "",
                "image":       e.get("Image"),
            })

    first_scene = scenes_raw[0]["SceneID"] if scenes_raw else None

    return {
        "first_scene": first_scene,
        "scenes": scenes,
    }


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"파일 없음: {EXCEL_PATH}")
        sys.exit(1)

    print(f"읽는 중: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    ignored_sheets = [name for name in wb.sheetnames if is_ignored_name(name)]
    if ignored_sheets:
        print(f"무시된 시트: {', '.join(ignored_sheets)}")

    data = build_game_data(wb)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json_str = json.dumps(data, ensure_ascii=False, indent=2)
        f.write(f"window.GAME_DATA = {json_str};\n")

    scene_count = len(data["scenes"])
    print(f"완료: {OUTPUT_PATH}")
    print(f"  씬 {scene_count}개 변환됨 / 첫 씬: {data['first_scene']}")


if __name__ == "__main__":
    main()

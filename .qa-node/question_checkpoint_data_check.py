import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TABLE_DIR = ROOT / "game" / "data" / "tables"
GENERATED_DIR = ROOT / "content" / "generated"
TABLE_KEYS = {
    "meta.json": "first_scene",
    "characters.json": "characters",
    "character_emotions.json": "character_emotions",
    "choice_groups.json": "choice_groups",
    "conditions.json": "conditions",
    "evidence_categories.json": "evidence_categories",
    "questions.json": "questions",
    "question_answers.json": "question_answers",
    "state_descriptors.json": "state_descriptors",
    "gauges.json": "gauges",
    "gauge_states.json": "gauge_states",
    "effects.json": "effects",
    "scenes.json": "scenes",
}


def load_bundle():
    text = (ROOT / "game" / "data" / "game_data.js").read_text(encoding="utf-8-sig").strip()
    text = re.sub(r"^window\.GAME_DATA\s*=\s*", "", text, count=1)
    return json.loads(re.sub(r";\s*$", "", text))


def main():
    bundle = load_bundle()
    table_payload = {}
    for filename, key in TABLE_KEYS.items():
        value = json.loads((TABLE_DIR / filename).read_text(encoding="utf-8-sig"))
        table_payload[key] = value["first_scene"] if filename == "meta.json" else value
    assert set(bundle) == set(table_payload) and bundle == table_payload

    answers = bundle["question_answers"]
    checkpoints = [scene for scene in bundle["scenes"].values() if scene.get("forced_question_ids")]
    assert len(answers) == 30 and len(checkpoints) == 5
    assert {answer["next_type"] for answer in answers} == {"Resume"}

    workbook = load_workbook(GENERATED_DIR / "script.generated.xlsx", read_only=True, data_only=True)
    answer_sheet = workbook["QuestionAnswerTable"]
    answer_values = answer_sheet.iter_rows(values_only=True)
    answer_headers = list(next(answer_values))
    answer_rows = list(answer_values)
    assert len(answer_headers) == 10 and len(answer_rows) == 30
    assert {row[answer_headers.index("NextType")] for row in answer_rows} == {"Resume"}
    question_sheet = workbook["QuestionTable"]
    question_values = question_sheet.iter_rows(values_only=True)
    question_headers = list(next(question_values))
    question_rows = list(question_values)
    visible_index = question_headers.index("VisibleConditionGroupIDs")
    assert len(question_rows) == 10 and all(question[visible_index] for question in question_rows)
    workbook.close()

    delimited_dir = GENERATED_DIR / "script.generated_delimited"
    delimited_files = sorted(delimited_dir.glob("*.csv")) + sorted(delimited_dir.glob("*.tsv"))
    for path in delimited_files:
        delimiter = "\t" if path.suffix == ".tsv" else ","
        with path.open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle, delimiter=delimiter))
        assert rows and all(len(row) == len(rows[0]) for row in rows[1:]), path
    for suffix in ("csv", "tsv"):
        path = delimited_dir / f"QuestionAnswerTable.{suffix}"
        with path.open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle, delimiter="\t" if suffix == "tsv" else ","))
        assert len(rows) == 31 and len(rows[0]) == 10
        next_type_index = rows[0].index("NextType")
        assert {row[next_type_index] for row in rows[1:]} == {"Resume"}

    print(json.dumps({
        "ok": True,
        "bundleKeys": len(bundle),
        "answers": len(answers),
        "checkpointScenes": len(checkpoints),
        "xlsx": {"answerRows": len(answer_rows), "answerColumns": len(answer_headers), "questionRows": len(question_rows)},
        "delimitedFiles": len(delimited_files),
        "questionAnswerDelimited": "30 rows x 10 columns",
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
